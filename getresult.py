from tkinter import *
from tkinter import ttk, messagebox
from PIL import Image,ImageTk
from tkcalendar import DateEntry
from openpyxl import *
from openpyxl.styles import Font,Border,Side,PatternFill,Alignment

rowinp = 5
r = 0

class Resultclass():
    def __init__(self,root):
        self.root=root
        root.title("TEACHER PORTAL")
        root.geometry('1200x650+100+10')
        self.root.resizable(False, False)
        self.root.config(bg="white")
        root.wm_iconbitmap("images//icon.ico")
        self.root.focus_force()
        self.Llbl = Frame(self.root, relief=RIDGE, bg="white")
        self.Llbl.place(x=0, y=40, height=690, width=700)
        self.Rlbl = Frame(self.root, relief=RIDGE, bg="white")
        self.Rlbl.place(x=700, y=40, height=690, width=500)
        Label(self.root, text="ADD STUDENT RESULT", font=("goudy ol style", 20, "bold"), bg="orange",fg="#262626").place(x=0, y=0, relwidth=1, height=40)

        bg_image = Image.open("images\\result.jpg")
        bg_image = bg_image.resize((580, 620), Image.Resampling.LANCZOS)
        self.bg_image = ImageTk.PhotoImage(bg_image)
        Label(self.Rlbl, image=self.bg_image).place(x=6, y=10, width=490, height=590)
        Label(self.Llbl, text="Roll no", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5,y=30)
        Label(self.Llbl, text="Name", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360, y=30)
        Label(self.Llbl, text="Registration no", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5,y=70)
        Label(self.Llbl, text="Fathers name", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5,y=110)
        Label(self.Llbl, text="Mothers Name", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360,y=70)
        Label(self.Llbl, text="D.O.B", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5, y=150)
        Label(self.Llbl, text="Name of exam", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360,y=110)
        Label(self.Llbl, text="SESSION", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360,y=150)

        Label(self.Llbl, text="Obtained Marks", font=("goudy ol style", 18, "bold"), bg="white", fg="blue").place(x=260,y=180)

        Label(self.Llbl, text="Phy-TH", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5, y=220)
        Label(self.Llbl, text="Phy-PR", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360, y=220)
        Label(self.Llbl, text="Chem-TH", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5,y=260)
        Label(self.Llbl, text="Chem-PR", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360,y=260)
        Label(self.Llbl, text="Maths", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5,y=300)
        Label(self.Llbl, text="English", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360,y=300)
        Label(self.Llbl, text="CS-PR", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=5,y=340)
        Label(self.Llbl, text="CS-TH", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=360,y=340)

        self.var_rollno = StringVar()
        self.var_name = StringVar()
        self.var_regno=StringVar()
        self.var_Fname = StringVar()
        self.var_Mname = StringVar()
        self.var_name_exam = StringVar()
        self.var_session = StringVar()
        self.var_phy_TH = StringVar()
        self.var_phy_PR = StringVar()
        self.var_chem_TH = StringVar()
        self.var_chem_PR = StringVar()
        self.var_maths = StringVar()
        self.var_cs_PR = StringVar()
        self.var_cs_TH = StringVar()
        self.var_english = StringVar()

        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_rollno,bg="light yellow", width=20).place(x=140, y=30)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_name, bg="light yellow",width=20).place(x=495, y=30)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_regno, bg="light yellow",width=20).place(x=140, y=70)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_Fname, bg="light yellow",width=20).place(x=140, y=110)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_Mname, bg="light yellow",width=20).place(x=495, y=70)
        self.cal = DateEntry(self.Llbl, selectmode="day")
        self.cal.place(x=140, y=150)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_name_exam, bg="light yellow",width=20).place(x=495, y=110)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_session, bg="light yellow",width=20).place(x=495, y=150)

        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_phy_TH, bg="light yellow",width=20).place(x=140, y=220)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_phy_PR, bg="light yellow",width=20).place(x=495, y=220)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_chem_TH, bg="light yellow",width=20).place(x=140, y=260)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_chem_PR, bg="light yellow",width=20).place(x=495, y=260)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_maths, bg="light yellow",width=20).place(x=140, y=300)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_english, bg="light yellow",width=20).place(x=495, y=300)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_cs_PR, bg="light yellow",width=20).place(x=140, y=340)
        Entry(self.Llbl, font=("times new roman", 15, "bold"), textvariable=self.var_cs_TH, bg="light yellow",width=20).place(x=495, y=340)

        Label(self.Llbl, text="OUT OF", font=("goudy ol style", 18, "bold"), bg="white", fg="blue").place(x=260,y=380)

        Label(self.Llbl, text="Phy-TH", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=45, y=420)
        Label(self.Llbl, text="Phy-PR", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=240,y=420)
        Label(self.Llbl, text="Chem-TH", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=420, y=420)
        Label(self.Llbl, text="Chem-PR", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=45,y=460)
        Label(self.Llbl, text="Maths", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=240, y=460)
        Label(self.Llbl, text="English", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=420, y=460)
        Label(self.Llbl, text="CS-PR", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=45, y=500)
        Label(self.Llbl, text="CS-TH", font=("times new roman", 15, "bold"), bg="white", fg="black").place(x=240, y=500)

        self.phy_TH = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.phy_TH["values"] = ("SELECT", "35", '40', "80")
        self.phy_TH.place(x=140, y=420, width=90, height=25)
        self.phy_TH.current(0)

        self.phy_PR = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.phy_PR["values"] = ("SELECT", "30", '25', "20")
        self.phy_PR.place(x=320, y=420, width=90, height=25)
        self.phy_PR.current(0)

        self.chem_TH = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.chem_TH["values"] = ("SELECT", "35", '40', "80")
        self.chem_TH.place(x=520, y=420, width=90, height=25)
        self.chem_TH.current(0)

        self.chem_PR = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.chem_PR["values"] = ("SELECT", "30", '25', "20")
        self.chem_PR.place(x=140, y=460, width=90, height=25)
        self.chem_PR.current(0)

        self.maths = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.maths["values"] = ("SELECT", "35", '40', "80")
        self.maths.place(x=320, y=460, width=90, height=25)
        self.maths.current(0)

        self.english = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.english["values"] = ("SELECT", "35", '40', "80")
        self.english.place(x=520, y=460, width=90, height=25)
        self.english.current(0)

        self.cs_PR = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.cs_PR["values"] = ("SELECT", "35", '40', "80")
        self.cs_PR.place(x=140, y=500, width=90, height=25)
        self.cs_PR.current(0)

        self.cs_TH = ttk.Combobox(self.Llbl, font=("times new roman", 10), justify=CENTER)
        self.cs_TH["values"] = ("SELECT", "35", '40', "80")
        self.cs_TH.place(x=320, y=500, width=90, height=25)
        self.cs_TH.current(0)

        self.btn_update = Button(self.Llbl, text="SAVE", font=("goudy old style", 15, "bold"), bg="#f44336",fg="white", cursor="hand2", command=self.add)
        self.btn_update.place(x=220, y=550, width=110, height=40)
        self.btn_delete = Button(self.Llbl, text="CLEAR", font=("goudy old style", 15, "bold"), bg="#0b5377",fg="white", cursor="hand2",command=self.clear)
        self.btn_delete.place(x=340, y=550, width=110, height=40)

    def clear(self):
        self.var_rollno.set("")
        self.var_name.set("")
        self.var_regno.set("")
        self.var_Fname.set("")
        self.var_Mname.set("")
        self.var_phy_TH.set("")
        self.var_phy_PR.set("")
        self.var_chem_TH.set("")
        self.var_chem_PR.set("")
        self.var_maths.set("")
        self.var_cs_PR.set("")
        self.var_cs_TH.set("")
        self.var_english.set("")

    def add(self):
        global rowinp,r
        eng, maths, phy_th, phy_pr, chem_th, chem_pr, cs_th, cs_pr = 4, 9, 14, 19, 22, 27, 30, 35

        if self.var_rollno.get()=="" :#or self.var_regno.get()=="" or self.var_name.get()=="" or self.var_Fname.get()==""=="" or self.var_Mname.get()=="" or self.cal.get_date()=="" or self.var_phy_TH.get()=="" or self.var_phy_PR.get()=="" or self.var_chem_TH.get()=="" or self.var_chem_PR.get()=="" or self.var_maths.get()=="" or self.var_cs_PR.get()=="" or self.var_cs_TH.get()=="" or self.var_english.get()=="" or self.var_phy_TH.get()=="SELECT" or self.phy_PR.get()=="SELECT" or self.var_chem_TH.get()=="select" or self.var_chem_PR.get()=="SELECT" or self.var_maths.get()=="SELECT" or self.var_english.get()=="SELECT" or self.var_cs_TH.get()=="SELECT" or self.var_cs_PR.get()=="SELECT" or self.var_session.get()=="":
                messagebox.showerror("ERROR", 'ALL FIELDS ARE REQUIRED \n PLEASE FILL ALL DETAILS')
        else:
            try:
                wb = load_workbook("studata2.xlsx")
                if self.var_session.get() not in wb.sheetnames:
                    ws=wb.create_sheet(f"{self.var_session.get()}")
                    ws.append(["ROLL NO.","NAME","REGISTRATION NO","KENDRIYA VIDHYALAYA NO.1 AFS GURUGRAM "])
                    ws.append(["","","","ENGLISH","","","","","MATHS","","","","","PHYSICS","","","","","","","","CHEMISTRY","","","","","","","","CS","","","","","","",""])
                    ws.append(["","","", "","","","","", "","","","","", "PHYSICS-TH","","","","" ,"PHYSICS-PR","","", "CHEMISTRY-TH", "", "", "","","CHEMISTRY-PR","","","CS-TH", "","", "", "","CS-PR"])
                    ws.cell(row=4, column=1).value = "Exam name->"
                    ws.merge_cells("D1:AK1")
                    ws.merge_cells("A1:A3")
                    ws.merge_cells("A4:C4")
                    ws.merge_cells("B1:B3")
                    ws.merge_cells("C1:C3")

                    ws.merge_cells("D2:H3")
                    ws.merge_cells("I2:M3")
                    ws.merge_cells("N2:U2")
                    ws.merge_cells("V2:AC2")
                    ws.merge_cells("AD2:AK2")

                    ws.merge_cells("N3:R3")
                    ws.merge_cells("S3:U3")
                    ws.merge_cells("V3:Z3")
                    ws.merge_cells("AA3:AC3")
                    ws.merge_cells("AD3:AH3")
                    ws.merge_cells("AI3:AK3")
                    for row in ws.iter_rows(min_row=4, min_col=14, max_row=4, max_col=21):
                        for cell in row:
                            if self.var_name_exam != cell.value:
                                while r < 1:
                                    ws.append([self.var_rollno.get(), self.var_name.get(), self.var_regno.get()])
                                    r += 1
                                rowinp=5
                                while eng<9:
                                    if ws.cell(row=4,column=eng).value is None:
                                        ws.cell(row=4, column=eng).value=self.var_name_exam.get()
                                        ws.cell(row=rowinp,column=eng).value=self.var_english.get()
                                        eng += 1
                                    break

                                while maths<14:
                                    if ws.cell(row=4,column=maths).value is None:
                                        ws.cell(row=4, column=maths).value=self.var_name_exam.get()
                                        ws.cell(row=rowinp, column=maths).value = self.var_maths.get()
                                        maths += 1
                                        break

                                while phy_th<19:
                                    if ws.cell(row=4,column=phy_th).value is None:
                                        ws.cell(row=4, column=phy_th).value=self.var_name_exam.get()
                                        ws.cell(row=rowinp, column=phy_th).value = self.var_phy_TH.get()
                                        phy_th += 1
                                        break

                                while chem_th<27:
                                    if ws.cell(row=4,column=chem_th).value is None:
                                        ws.cell(row=4, column=chem_th).value=self.var_name_exam.get()
                                        ws.cell(row=rowinp, column=chem_th).value = self.var_chem_TH.get()
                                        chem_th += 1
                                        break

                                while cs_th<35:
                                    if ws.cell(row=4,column=cs_th).value is None:
                                        ws.cell(row=4, column=cs_th).value=self.var_name_exam.get()
                                        ws.cell(row=rowinp, column=cs_th).value = self.var_cs_TH.get()
                                        cs_th += 1
                                        break

                                    if cs_th==34:
                                        rowinp+=1
                            else:
                                while eng < 9:
                                    if ws.cell(row=rowinp, column=eng).value is None:
                                        ws.cell(row=rowinp, column=eng).value=self.english
                                        eng+=1
                                        break

                                while maths < 9:
                                    if ws.cell(row=rowinp, column=maths).value is None:
                                        ws.cell(row=rowinp, column=maths).value=self.english
                                        maths += 1
                                        break

                                while phy_th < 9:
                                    if ws.cell(row=rowinp, column=phy_th).value is None:
                                        ws.cell(row=rowinp, column=phy_th).value=self.english
                                        phy_th += 1
                                        break
                                while chem_th < 9:
                                    if ws.cell(row=rowinp, column=chem_th).value is None:
                                        ws.cell(row=rowinp, column=chem_th).value=self.english
                                        chem_th+=1
                                        break

                                while cs_th < 9:
                                    if ws.cell(row=rowinp, column=cs_th).value is None:
                                        ws.cell(row=rowinp, column=cs_th).value=self.english
                                        cs_th+=1
                                        break
                                    if cs_th==34:
                                        rowinp+=1

                    thin = Side(border_style="medium", color="000000")
                    thick = Side(border_style="thick", color="000000")
                    for i in range(1,4):
                        for j in range(1,4):
                            ws.cell(row=i, column=j).font = Font(size=11, bold=True)
                            ws.cell(row=i, column=j).fill = PatternFill(fill_type="solid", fgColor="00FFFF00")
                            ws.cell(row=i, column=j).border = Border(top=thick, left=thick, right=thick, bottom=thick)
                    ws.cell(row=1, column=4).font = Font(size=15, bold=True)
                    ws.cell(row=4, column=1).font = Font(size=15, bold=True)
                    ws.cell(row=1, column=4).fill = PatternFill(fill_type="solid", fgColor="0033CCCC")
                    ws.cell(row=4, column=1).fill = PatternFill(fill_type="solid", fgColor="0000FFFF")
                    for i in range(4,38):
                        ws.cell(row=1, column=i).border = Border(top=thick, left=thick, right=thick, bottom=thick)
                        ws.cell(row=2, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        ws.cell(row=2,column=i).alignment = Alignment(horizontal='center')
                        ws.cell(row=3,column=i).alignment = Alignment(horizontal='center')
                    for i in range(4,14):
                        ws.cell(row=3, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        ws.cell(row=2, column=i).font = Font(size=11, bold=True)
                    for i in range(14,38):
                        ws.cell(row=2, column=i).font = Font(size=11, bold=True)
                        ws.cell(row=3, column=i).font = Font(size=11, bold=True)
                        ws.cell(row=3, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    for i in range(1,38):
                        ws.cell(row=4, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    ws.cell(row=4, column=3).border = Border(right=thin,bottom=thin)
                    ws.column_dimensions['A'].height = 3
                    ws.column_dimensions['B'].height = 3
                    ws.column_dimensions['C'].width = 19
                    ws["D1"].alignment = Alignment(horizontal='center')
                    ws["A4"].alignment = Alignment(horizontal='center')
                    for i in ("D","V"):
                        ws[i+"2"].fill = PatternFill(fill_type="solid", fgColor="00FF9900")
                    for i in ("I", "AD"):
                        ws[i+"2"].fill = PatternFill(fill_type="solid", fgColor="0099CC00")
                        ws[i+"3"].fill = PatternFill(fill_type="solid", fgColor="00FF9900")
                    ws["N3"].fill=PatternFill(fill_type="solid", fgColor="00FF9900")
                    ws["S3"].fill = PatternFill(fill_type="solid", fgColor="0099CC00")
                    ws["V3"].fill = PatternFill(fill_type="solid", fgColor="0099CC00")
                    for i in ("D","E","F","G","H","S","T","U","AD","AE","AF","AG","AH"):
                        ws[i+"4"].fill = PatternFill(fill_type="solid", fgColor="009999FF")
                    for i in ("N","O","P","Q","R","AA","AB","AC","AI","AJ","AK"):
                        ws[i+"4"].fill = PatternFill(fill_type="solid", fgColor="0099CCFF")
                    messagebox.showinfo("Success", "Marks Added Successfully")
                else:
                    wb = load_workbook("studata2.xlsx")
                    ws=wb[self.var_session.get()]
                    while eng<9:
                        if ws.cell(row=4,column=eng).value is None:
                            ws.cell(row=4, column=eng).value=self.var_name_exam.get()
                            ws.cell(row=rowinp,column=eng).value=self.var_english.get()
                            break
                        eng+=1
                    while maths<14:
                        if ws.cell(row=4,column=maths).value is None:
                            ws.cell(row=4, column=maths).value=self.var_name_exam.get()
                            ws.cell(row=rowinp, column=maths).value = self.var_maths.get()
                            break
                        maths+=1
                    while phy_th<19:
                        if ws.cell(row=4,column=phy_th).value is None:
                            ws.cell(row=4, column=phy_th).value=self.var_name_exam.get()
                            ws.cell(row=rowinp, column=phy_th).value = self.var_phy_TH.get()
                            break
                        phy_th+=1
                    while chem_th<27:
                        if ws.cell(row=4,column=chem_th).value is None:
                            ws.cell(row=4, column=chem_th).value=self.var_name_exam.get()
                            ws.cell(row=rowinp, column=chem_th).value = self.var_chem_TH.get()
                            break
                        chem_th+=1
                    while cs_th<35:
                        if ws.cell(row=4,column=cs_th).value is None:
                            ws.cell(row=4, column=cs_th).value=self.var_name_exam.get()
                            ws.cell(row=rowinp, column=cs_th).value = self.var_cs_TH.get()
                            break
                        cs_th+=1
                        if cs_th==34:
                            rowinp+=1
                    messagebox.showinfo("Success", "Marks Added Successfully")
                wb.save("studata2.xlsx")
            except Exception as es:
                messagebox.showerror("error", f"Error due to {str(es)}", parent=self.root)

if __name__=="__main__":
    root=Tk()
    obj=Resultclass(root)
    root.mainloop()